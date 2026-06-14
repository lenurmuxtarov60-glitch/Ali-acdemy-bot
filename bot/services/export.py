import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from bot.models import Group, Student, get_session
from bot.services.attendance import AttendanceService
from bot.utils.helpers import date_range


class ExportService:

    @staticmethod
    def export_excel(period: str = "monthly", group_id: int | None = None) -> io.BytesIO:
        start, end = date_range(period)
        wb = Workbook()
        ws = wb.active
        ws.title = "Davomat"

        headers = ["№", "F.I.O", "Guruh", "Kurs", "Keldi", "Kelmadi", "Kechikdi", "Foiz (%)"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        with get_session() as session:
            if group_id:
                groups = [session.query(Group).filter_by(id=group_id).first()]
            else:
                groups = session.query(Group).filter_by(is_active=True).all()

            row = 2
            num = 1
            for group in groups:
                if not group:
                    continue
                report = AttendanceService.get_group_attendance_report(group.id, start, end)
                for item in report:
                    s = item["student"]
                    ws.cell(row=row, column=1, value=num)
                    ws.cell(row=row, column=2, value=s.full_name)
                    ws.cell(row=row, column=3, value=group.name)
                    ws.cell(row=row, column=4, value=group.course)
                    ws.cell(row=row, column=5, value=item["present"])
                    ws.cell(row=row, column=6, value=item["absent"])
                    ws.cell(row=row, column=7, value=item["late"])
                    ws.cell(row=row, column=8, value=f"{item['percentage']:.1f}")
                    row += 1
                    num += 1

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_pdf(period: str = "monthly", group_id: int | None = None) -> io.BytesIO:
        start, end = date_range(period)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=16,
            spaceAfter=12,
            alignment=1,
        )

        period_names = {"daily": "Kunlik", "weekly": "Haftalik", "monthly": "Oylik"}
        title = f"{period_names.get(period, 'Davomat')} hisobot"
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(
            f"Davr: {start.strftime('%d.%m.%Y')} — {end.strftime('%d.%m.%Y')}",
            styles["Normal"],
        ))
        elements.append(Spacer(1, 20))

        data = [["F.I.O", "Guruh", "Keldi", "Kelmadi", "Kechikdi", "Foiz"]]

        with get_session() as session:
            if group_id:
                groups = [session.query(Group).filter_by(id=group_id).first()]
            else:
                groups = session.query(Group).filter_by(is_active=True).all()

            for group in groups:
                if not group:
                    continue
                report = AttendanceService.get_group_attendance_report(group.id, start, end)
                for item in report:
                    s = item["student"]
                    data.append([
                        s.full_name,
                        group.name,
                        str(item["present"]),
                        str(item["absent"]),
                        str(item["late"]),
                        f"{item['percentage']:.1f}%",
                    ])

        table = Table(data, colWidths=[120, 80, 50, 60, 60, 50])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        return buffer
